from Array import Array
from openpyxl import load_workbook
from scipy.optimize import fsolve


def light():
    ambient_water = float(input("Ambient water vapor mole fraction (ppt): "))
    ambient_carbon = float(input("Ambient carbon dioxide mole fraction (ppm): "))
    ambient_temperature = float(input("Ambient temperature (K): "))
    total_intensity = float(input("Total incident light intensity (W/m^2): "))
    blue_intensity = float(input("Incident blue light intensity (W/m^2): "))

    length = 1
    width = 1

    array = Array(length, width)

    neighbors_carbon = []
    neighbors_water = []

    for i in range(length):
        for unit in array[i]:
            unit.set_ambient_water(ambient_water)
            unit.set_ambient_temperature(ambient_temperature)
            unit.set_ambient_carbon(ambient_carbon)
            unit.set_total_intensity(0)
            unit.set_blue_intensity(0)
            unit.set_absorbed_intensity()

    out_file = load_workbook('C:\\Users\\Undergrunt\\Box\\Model\\Test.xlsx')
    ws = out_file.create_sheet("Test")

    count = 0
    while count < 150:
        for i in range(length):
            for unit in array[i]:
                neighbors_carbon.clear()
                neighbors_water.clear()
                neighbors_carbon.append(unit.get_carbon_dioxide())
                neighbors_water.append(unit.epid.get_epid_water_potential())
                average_carbon = sum(neighbors_carbon) / len(neighbors_carbon)
                average_water = sum(neighbors_water) / len(neighbors_water)
                unit.calculate_next(average_water)
                x, y = fsolve(unit.solve_for_temperature, (296, 30))
                unit.set_temperature(x)
                unit.set_es_water_vapor(y)
                unit.set_es_water_potential(unit.calculate_es_water_potential(ambient_water))
                unit.set_carbon_dioxide(unit.calculate_carbon_dioxide(0, ambient_carbon, average_carbon))
                unit.guard.set_guard_signal(unit.guard.calculate_guard_signal(0, 0,
                                                                              unit.get_carbon_dioxide()))
                unit.epid.set_epid_signal(unit.epid.calculate_epid_signal(0, unit.get_carbon_dioxide()))
                unit.guard.set_guard_water_potential(unit.guard.calculate_guard_water_potential(unit.get_temperature()))
                unit.epid.set_epid_water_potential(unit.epid.calculate_epid_water_potential(unit.get_temperature()))
                unit.set_pore_water_vapor(unit.calculate_pore_water_vapor(ambient_water))
                unit.set_pore_water_potential(unit.calculate_pore_water_potential())

                ws.cell(row=count+1, column=1).value = unit.get_conductance()
                ws.cell(row=count+1, column=2).value = unit.get_temperature()
                ws.cell(row=count+1, column=3).value = unit.guard.get_guard_pressure()
                ws.cell(row=count+1, column=4).value = unit.epid.get_epid_pressure()
                ws.cell(row=count+1, column=5).value = unit.guard.get_guard_water_potential()
                ws.cell(row=count+1, column=6).value = unit.epid.get_epid_water_potential()
                ws.cell(row=count+1, column=7).value = unit.guard.get_guard_signal() * 8.314 * unit.get_temperature()
                ws.cell(row=count+1, column=8).value = unit.epid.get_epid_signal() * 8.314 * unit.get_temperature()
                ws.cell(row=count+1, column=9).value = unit.get_es_water_vapor()
                ws.cell(row=count+1, column=10).value = unit.get_pore_water_potential()
                ws.cell(row=count+1, column=11).value = unit.get_pore_water_vapor()
                ws.cell(row=count+1, column=12).value = unit.get_es_water_potential()
                ws.cell(row=count+1, column=13).value = unit.get_carbon_dioxide()

        count += 1

    for i in range(length):
        for unit in array[i]:
            unit.set_total_intensity(total_intensity)
            unit.set_absorbed_intensity()

    while count < 160:
        for i in range(length):
            for unit in array[i]:
                neighbors_carbon.clear()
                neighbors_water.clear()
                neighbors_carbon.append(unit.get_carbon_dioxide())
                neighbors_water.append(unit.epid.get_epid_water_potential())
                average_carbon = sum(neighbors_carbon) / len(neighbors_carbon)
                average_water = sum(neighbors_water) / len(neighbors_water)
                unit.calculate_next(average_water)
                x, y = fsolve(unit.solve_for_temperature, (296, 30))
                unit.set_temperature(x)
                unit.set_es_water_vapor(y)
                unit.set_es_water_potential(unit.calculate_es_water_potential(ambient_water))
                unit.set_carbon_dioxide(unit.calculate_carbon_dioxide(0, ambient_carbon, average_carbon))
                unit.guard.set_guard_signal(unit.guard.calculate_guard_signal(0, 0,
                                                                              ws.cell(row=count-5, column=13).value))
                unit.epid.set_epid_signal(unit.epid.calculate_epid_signal(0,  ws.cell(row=count-5, column=13).value))
                unit.guard.set_guard_water_potential(unit.guard.calculate_guard_water_potential(unit.get_temperature()))
                unit.epid.set_epid_water_potential(unit.epid.calculate_epid_water_potential(unit.get_temperature()))
                unit.set_pore_water_vapor(unit.calculate_pore_water_vapor(ambient_water))
                unit.set_pore_water_potential(unit.calculate_pore_water_potential())

                ws.cell(row=count+1, column=1).value = unit.get_conductance()
                ws.cell(row=count+1, column=2).value = unit.get_temperature()
                ws.cell(row=count+1, column=3).value = unit.guard.get_guard_pressure()
                ws.cell(row=count+1, column=4).value = unit.epid.get_epid_pressure()
                ws.cell(row=count+1, column=5).value = unit.guard.get_guard_water_potential()
                ws.cell(row=count+1, column=6).value = unit.epid.get_epid_water_potential()
                ws.cell(row=count+1, column=7).value = unit.guard.get_guard_signal() * 8.314 * unit.get_temperature()
                ws.cell(row=count+1, column=8).value = unit.epid.get_epid_signal() * 8.314 * unit.get_temperature()
                ws.cell(row=count+1, column=9).value = unit.get_es_water_vapor()
                ws.cell(row=count+1, column=10).value = unit.get_pore_water_potential()
                ws.cell(row=count+1, column=11).value = unit.get_pore_water_vapor()
                ws.cell(row=count+1, column=12).value = unit.get_es_water_potential()
                ws.cell(row=count+1, column=13).value = unit.get_carbon_dioxide()

        count += 1

    while count < 300:
        for i in range(length):
            for unit in array[i]:
                neighbors_carbon.clear()
                neighbors_water.clear()
                neighbors_carbon.append(unit.get_carbon_dioxide())
                neighbors_water.append(unit.epid.get_epid_water_potential())
                average_carbon = sum(neighbors_carbon) / len(neighbors_carbon)
                average_water = sum(neighbors_water) / len(neighbors_water)
                unit.calculate_next(average_water)
                x, y = fsolve(unit.solve_for_temperature, (296, 30))
                unit.set_temperature(x)
                unit.set_es_water_vapor(y)
                unit.set_es_water_potential(unit.calculate_es_water_potential(ambient_water))
                unit.set_carbon_dioxide(unit.calculate_carbon_dioxide(total_intensity, ambient_carbon, average_carbon))
                unit.guard.set_guard_signal(unit.guard.calculate_guard_signal(0, total_intensity,
                                                                              ws.cell(row=count-5, column=13).value))
                unit.epid.set_epid_signal(unit.epid.calculate_epid_signal(total_intensity, ws.cell(row=count-5,
                                                                                                   column=13).value))
                unit.guard.set_guard_water_potential(unit.guard.calculate_guard_water_potential(unit.get_temperature()))
                unit.epid.set_epid_water_potential(unit.epid.calculate_epid_water_potential(unit.get_temperature()))
                unit.set_pore_water_vapor(unit.calculate_pore_water_vapor(ambient_water))
                unit.set_pore_water_potential(unit.calculate_pore_water_potential())

                ws.cell(row=count + 1, column=1).value = unit.get_conductance()
                ws.cell(row=count + 1, column=2).value = unit.get_temperature()
                ws.cell(row=count + 1, column=3).value = unit.guard.get_guard_pressure()
                ws.cell(row=count + 1, column=4).value = unit.epid.get_epid_pressure()
                ws.cell(row=count + 1, column=5).value = unit.guard.get_guard_water_potential()
                ws.cell(row=count + 1, column=6).value = unit.epid.get_epid_water_potential()
                ws.cell(row=count + 1, column=7).value = unit.guard.get_guard_signal() * 8.314 * unit.get_temperature()
                ws.cell(row=count + 1, column=8).value = unit.epid.get_epid_signal() * 8.314 * unit.get_temperature()
                ws.cell(row=count + 1, column=9).value = unit.get_es_water_vapor()
                ws.cell(row=count + 1, column=10).value = unit.get_pore_water_potential()
                ws.cell(row=count + 1, column=11).value = unit.get_pore_water_vapor()
                ws.cell(row=count + 1, column=12).value = unit.get_es_water_potential()
                ws.cell(row=count + 1, column=13).value = unit.get_carbon_dioxide()

        count += 1

    for i in range(length):
        for unit in array[i]:
            unit.set_total_intensity(total_intensity)
            unit.set_blue_intensity(blue_intensity)
            unit.set_absorbed_intensity()

        while count < 450:
            for i in range(length):
                for unit in array[i]:
                    neighbors_carbon.clear()
                    neighbors_water.clear()
                    neighbors_carbon.append(unit.get_carbon_dioxide())
                    neighbors_water.append(unit.epid.get_epid_water_potential())
                    average_carbon = sum(neighbors_carbon) / len(neighbors_carbon)
                    average_water = sum(neighbors_water) / len(neighbors_water)
                    unit.calculate_next(average_water)
                    x, y = fsolve(unit.solve_for_temperature, (296, 30))
                    unit.set_temperature(x)
                    unit.set_es_water_vapor(y)
                    unit.set_es_water_potential(unit.calculate_es_water_potential(ambient_water))
                    unit.set_carbon_dioxide(
                        unit.calculate_carbon_dioxide(total_intensity, ambient_carbon, average_carbon))
                    unit.guard.set_guard_signal(unit.guard.calculate_guard_signal(blue_intensity, total_intensity,
                                                                                  ws.cell(row=count - 5,
                                                                                          column=13).value))
                    unit.epid.set_epid_signal(unit.epid.calculate_epid_signal(total_intensity, ws.cell(row=count - 5,
                                                                                                       column=13).value))
                    unit.guard.set_guard_water_potential(
                        unit.guard.calculate_guard_water_potential(unit.get_temperature()))
                    unit.epid.set_epid_water_potential(unit.epid.calculate_epid_water_potential(unit.get_temperature()))
                    unit.set_pore_water_vapor(unit.calculate_pore_water_vapor(ambient_water))
                    unit.set_pore_water_potential(unit.calculate_pore_water_potential())

                    ws.cell(row=count + 1, column=1).value = unit.get_conductance()
                    ws.cell(row=count + 1, column=2).value = unit.get_temperature()
                    ws.cell(row=count + 1, column=3).value = unit.guard.get_guard_pressure()
                    ws.cell(row=count + 1, column=4).value = unit.epid.get_epid_pressure()
                    ws.cell(row=count + 1, column=5).value = unit.guard.get_guard_water_potential()
                    ws.cell(row=count + 1, column=6).value = unit.epid.get_epid_water_potential()
                    ws.cell(row=count + 1,
                            column=7).value = unit.guard.get_guard_signal() * 8.314 * unit.get_temperature()
                    ws.cell(row=count + 1,
                            column=8).value = unit.epid.get_epid_signal() * 8.314 * unit.get_temperature()
                    ws.cell(row=count + 1, column=9).value = unit.get_es_water_vapor()
                    ws.cell(row=count + 1, column=10).value = unit.get_pore_water_potential()
                    ws.cell(row=count + 1, column=11).value = unit.get_pore_water_vapor()
                    ws.cell(row=count + 1, column=12).value = unit.get_es_water_potential()
                    ws.cell(row=count + 1, column=13).value = unit.get_carbon_dioxide()

            count += 1

    out_file.save('C:\\Users\\Undergrunt\\Box\\Model\\Test.xlsx')


def humidity():
    ambient_water = float(input("Ambient water vapor mole fraction (ppt): "))
    ambient_carbon = float(input("Ambient carbon dioxide mole fraction (ppm): "))
    ambient_temperature = float(input("Ambient temperature (K): "))
    total_intensity = float(input("Total incident light intensity (W/m^2): "))
    blue_intensity = float(input("Incident blue light intensity (W/m^2): "))

    length = 1
    width = 1

    array = Array(length, width)

    neighbors_carbon = []
    neighbors_water = []

    for i in range(length):
        for unit in array[i]:
            unit.set_ambient_water(ambient_water)
            unit.set_ambient_temperature(ambient_temperature)
            unit.set_ambient_carbon(ambient_carbon)
            unit.set_total_intensity(total_intensity)
            unit.set_blue_intensity(blue_intensity)
            unit.set_absorbed_intensity()

    out_file = load_workbook('C:\\Users\\Undergrunt\\Box\\Model\\Test.xlsx')
    ws = out_file.create_sheet("Test")

    count = 0
    while count < 150:
        for i in range(length):
            for unit in array[i]:
                neighbors_carbon.clear()
                neighbors_water.clear()
                neighbors_carbon.append(unit.get_carbon_dioxide())
                neighbors_water.append(unit.epid.get_epid_water_potential())
                average_carbon = sum(neighbors_carbon) / len(neighbors_carbon)
                average_water = sum(neighbors_water) / len(neighbors_water)
                unit.calculate_next(average_water)
                x, y = fsolve(unit.solve_for_temperature, (296, 30))
                unit.set_temperature(x)
                unit.set_es_water_vapor(y)
                unit.set_es_water_potential(unit.calculate_es_water_potential(ambient_water))
                unit.set_carbon_dioxide(unit.calculate_carbon_dioxide(total_intensity, ambient_carbon, average_carbon))
                unit.guard.set_guard_signal(unit.guard.calculate_guard_signal(blue_intensity, total_intensity,
                                                                              unit.get_carbon_dioxide()))
                unit.epid.set_epid_signal(unit.epid.calculate_epid_signal(total_intensity, unit.get_carbon_dioxide()))
                unit.guard.set_guard_water_potential(unit.guard.calculate_guard_water_potential(unit.get_temperature()))
                unit.epid.set_epid_water_potential(unit.epid.calculate_epid_water_potential(unit.get_temperature()))
                unit.set_pore_water_vapor(unit.calculate_pore_water_vapor(ambient_water))
                unit.set_pore_water_potential(unit.calculate_pore_water_potential())

                ws.cell(row=count+1, column=1).value = unit.get_conductance()
                ws.cell(row=count+1, column=2).value = unit.get_temperature()
                ws.cell(row=count+1, column=3).value = unit.guard.get_guard_pressure()
                ws.cell(row=count+1, column=4).value = unit.epid.get_epid_pressure()
                ws.cell(row=count+1, column=5).value = unit.guard.get_guard_water_potential()
                ws.cell(row=count+1, column=6).value = unit.epid.get_epid_water_potential()
                ws.cell(row=count+1, column=7).value = unit.guard.get_guard_signal() * 8.314 * unit.get_temperature()
                ws.cell(row=count+1, column=8).value = unit.epid.get_epid_signal() * 8.314 * unit.get_temperature()
                ws.cell(row=count+1, column=9).value = unit.get_es_water_vapor()
                ws.cell(row=count+1, column=10).value = unit.get_pore_water_potential()
                ws.cell(row=count+1, column=11).value = unit.get_pore_water_vapor()
                ws.cell(row=count+1, column=12).value = unit.get_es_water_potential()
                ws.cell(row=count+1, column=13).value = unit.get_carbon_dioxide()

        count += 1

    for i in range(length):
        for unit in array[i]:
            unit.set_total_intensity(total_intensity)
            unit.set_blue_intensity(blue_intensity)
            unit.set_absorbed_intensity()

    while count < 160:
        for i in range(length):
            for unit in array[i]:
                neighbors_carbon.clear()
                neighbors_water.clear()
                neighbors_carbon.append(unit.get_carbon_dioxide())
                neighbors_water.append(unit.epid.get_epid_water_potential())
                average_carbon = sum(neighbors_carbon) / len(neighbors_carbon)
                average_water = sum(neighbors_water) / len(neighbors_water)
                unit.calculate_next(average_water)
                x, y = fsolve(unit.solve_for_temperature, (296, 30))
                unit.set_temperature(x)
                unit.set_es_water_vapor(y)
                unit.set_es_water_potential(unit.calculate_es_water_potential(ambient_water - 10))
                unit.set_carbon_dioxide(unit.calculate_carbon_dioxide(total_intensity, ambient_carbon, average_carbon))
                unit.guard.set_guard_signal(unit.guard.calculate_guard_signal(blue_intensity, total_intensity,
                                                                              ws.cell(row=count-10, column=13).value))
                unit.epid.set_epid_signal(unit.epid.calculate_epid_signal(total_intensity,  ws.cell(row=count-10,
                                                                                                    column=13).value))
                unit.guard.set_guard_water_potential(unit.guard.calculate_guard_water_potential(unit.get_temperature()))
                unit.epid.set_epid_water_potential(unit.epid.calculate_epid_water_potential(unit.get_temperature()))
                unit.set_pore_water_vapor(unit.calculate_pore_water_vapor(ambient_water - 10))
                unit.set_pore_water_potential(unit.calculate_pore_water_potential())

                ws.cell(row=count+1, column=1).value = unit.get_conductance()
                ws.cell(row=count+1, column=2).value = unit.get_temperature()
                ws.cell(row=count+1, column=3).value = unit.guard.get_guard_pressure()
                ws.cell(row=count+1, column=4).value = unit.epid.get_epid_pressure()
                ws.cell(row=count+1, column=5).value = unit.guard.get_guard_water_potential()
                ws.cell(row=count+1, column=6).value = unit.epid.get_epid_water_potential()
                ws.cell(row=count+1, column=7).value = unit.guard.get_guard_signal() * 8.314 * unit.get_temperature()
                ws.cell(row=count+1, column=8).value = unit.epid.get_epid_signal() * 8.314 * unit.get_temperature()
                ws.cell(row=count+1, column=9).value = unit.get_es_water_vapor()
                ws.cell(row=count+1, column=10).value = unit.get_pore_water_potential()
                ws.cell(row=count+1, column=11).value = unit.get_pore_water_vapor()
                ws.cell(row=count+1, column=12).value = unit.get_es_water_potential()
                ws.cell(row=count+1, column=13).value = unit.get_carbon_dioxide()

        count += 1

    while count < 400:
        for i in range(length):
            for unit in array[i]:
                neighbors_carbon.clear()
                neighbors_water.clear()
                neighbors_carbon.append(unit.get_carbon_dioxide())
                neighbors_water.append(unit.epid.get_epid_water_potential())
                average_carbon = sum(neighbors_carbon) / len(neighbors_carbon)
                average_water = sum(neighbors_water) / len(neighbors_water)
                unit.calculate_next(average_water)
                x, y = fsolve(unit.solve_for_temperature, (296, 30))
                unit.set_temperature(x)
                unit.set_es_water_vapor(y)
                unit.set_es_water_potential(unit.calculate_es_water_potential(ambient_water - 10))
                unit.set_carbon_dioxide(unit.calculate_carbon_dioxide(total_intensity, ambient_carbon, average_carbon))
                unit.guard.set_guard_signal(unit.guard.calculate_guard_signal(blue_intensity, total_intensity,
                                                                              ws.cell(row=count-10, column=13).value))
                unit.epid.set_epid_signal(unit.epid.calculate_epid_signal(total_intensity, ws.cell(row=count-10,
                                                                                                   column=13).value))
                unit.guard.set_guard_water_potential(unit.guard.calculate_guard_water_potential(unit.get_temperature()))
                unit.epid.set_epid_water_potential(unit.epid.calculate_epid_water_potential(unit.get_temperature()))
                unit.set_pore_water_vapor(unit.calculate_pore_water_vapor(ambient_water - 10))
                unit.set_pore_water_potential(unit.calculate_pore_water_potential())

                ws.cell(row=count + 1, column=1).value = unit.get_conductance()
                ws.cell(row=count + 1, column=2).value = unit.get_temperature()
                ws.cell(row=count + 1, column=3).value = unit.guard.get_guard_pressure()
                ws.cell(row=count + 1, column=4).value = unit.epid.get_epid_pressure()
                ws.cell(row=count + 1, column=5).value = unit.guard.get_guard_water_potential()
                ws.cell(row=count + 1, column=6).value = unit.epid.get_epid_water_potential()
                ws.cell(row=count + 1, column=7).value = unit.guard.get_guard_signal() * 8.314 * unit.get_temperature()
                ws.cell(row=count + 1, column=8).value = unit.epid.get_epid_signal() * 8.314 * unit.get_temperature()
                ws.cell(row=count + 1, column=9).value = unit.get_es_water_vapor()
                ws.cell(row=count + 1, column=10).value = unit.get_pore_water_potential()
                ws.cell(row=count + 1, column=11).value = unit.get_pore_water_vapor()
                ws.cell(row=count + 1, column=12).value = unit.get_es_water_potential()
                ws.cell(row=count + 1, column=13).value = unit.get_carbon_dioxide()

        count += 1

    out_file.save('C:\\Users\\Undergrunt\\Box\\Model\\Test.xlsx')


light()