from Array import Array
from openpyxl import load_workbook
from scipy.optimize import fsolve


def array_main(ambient_water, ambient_carbon, ambient_temperature, total_intensity, blue_intensity, length, width):

    array1 = Array(length, width)
    array2 = Array(length, width)
    array3 = Array(length, width)
    array4 = Array(length, width)
    array5 = Array(length, width)
    array6 = Array(length, width)
    array7 = Array(length, width)

    array1.randomize()

    neighbors_carbon = []
    neighbors_water = []

    for i in range(length):
        for unit in array1[i]:
            unit.set_ambient_water(ambient_water)
            unit.set_ambient_temperature(ambient_temperature)
            unit.set_ambient_carbon(ambient_carbon)
            unit.set_total_intensity(0)
            unit.set_blue_intensity(0)
            unit.set_absorbed_intensity()

            array2[i][unit.get_col()].set_carbon_dioxide(unit.get_carbon_dioxide)
            array2[i][unit.get_col()].epid.set_epid_water_potential(unit.epid.get_epid_water_potential)

    temp_file = load_workbook('C:\\Users\\Undergrunt\\Box\\Model\\Temperature.xlsx')
    cond_file = load_workbook('C:\\Users\\Undergrunt\\Box\\Model\\Conductance.xlsx')

    count = 0
    while count < 20:
        ws1 = temp_file.create_sheet(str(count))
        ws2 = cond_file.create_sheet(str(count))

        for i in range(len(array1)):
            for unit in array1[i]:
                neighbors_carbon.clear()
                neighbors_water.clear()

                if i != 0 and unit.get_col() != 0 and i != length - 1 and unit.get_col() != width - 1:
                    neighbors_carbon.append(array2[i][unit.get_col() + 1].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[i][unit.get_col() - 1].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[i + 1][unit.get_col()].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[i - 1][unit.get_col()].get_carbon_dioxide()())

                    neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                elif i == 0 and unit.get_col() != 0 and unit.get_col() != width - 1:
                    neighbors_carbon.append(array2[i][unit.get_col() + 1].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[i][unit.get_col() - 1].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[i + 1][unit.get_col()].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[length - 1][unit.get_col()].get_carbon_dioxide()())

                    neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[length - 1][unit.get_col()].epid.get_epid_water_potential()())

                elif unit.get_col() == 0 and i != length - 1 and i != 0:
                    neighbors_carbon.append(array2[i][unit.get_col() + 1].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[i][width - 1].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[i + 1][unit.get_col()].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[i - 1][unit.get_col()].get_carbon_dioxide()())

                    neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                elif i == length - 1 and unit.get_col() != 0 and unit.get_col() != width - 1:
                    neighbors_carbon.append(array2[i][unit.get_col() + 1].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[i][unit.get_col() - 1].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[i - 1][unit.get_col()].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[0][unit.get_col()].get_carbon_dioxide()())

                    neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[0][unit.get_col()].epid.get_epid_water_potential()())

                elif unit.get_col() == width - 1 and i != 0 and i != length - 1:
                    neighbors_carbon.append(array2[i][unit.get_col() - 1].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[i][0].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[i + 1][unit.get_col()].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[i - 1][unit.get_col()].get_carbon_dioxide()())

                    neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                elif i == 0 and unit.get_col() == 0:
                    neighbors_carbon.append(array2[i][unit.get_col() + 1].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[i][width - 1].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[i + 1][unit.get_col()].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[length - 1][unit.get_col()].get_carbon_dioxide()())

                    neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[length - 1][unit.get_col()].epid.get_epid_water_potential()())

                elif i == 0 and unit.get_col() == width - 1:
                    neighbors_carbon.append(array2[i][unit.get_col() - 1].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[i][0].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[i + 1][width - 1].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[length - 1][width - 1].get_carbon_dioxide()())

                    neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[i + 1][width - 1].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[length - 1][width - 1].epid.get_epid_water_potential()())

                elif i == length - 1 and unit.get_col() == 0:
                    neighbors_carbon.append(array2[i][unit.get_col() + 1].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[i][width - 1].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[i - 1][unit.get_col()].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[0][unit.get_col()].get_carbon_dioxide()())

                    neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[1][unit.get_col()].epid.get_epid_water_potential()())

                elif i == length - 1 and unit.get_col() == width - 1:
                    neighbors_carbon.append(array2[i][unit.get_col() - 1].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[i][0].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[i - 1][unit.get_col()].get_carbon_dioxide()())
                    neighbors_carbon.append(array2[0][width - 1].get_carbon_dioxide()())

                    neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                    neighbors_water.append(array2[0][width - 1].epid.get_epid_water_potential()())

                average_carbon = sum(neighbors_carbon) / len(neighbors_carbon)
                average_water = sum(neighbors_water) / len(neighbors_water)
                unit.calculate_next(average_water)
                x, y = fsolve(unit.solve_for_temperature, (296, 30))
                unit.set_temperature(x)
                unit.set_es_water_vapor(y)
                unit.set_es_water_potential(unit.calculate_es_water_potential(ambient_water))
                unit.set_carbon_dioxide(unit.calculate_carbon_dioxide(0, ambient_carbon, average_carbon))
                unit.guard.set_guard_signal(unit.guard.calculate_guard_signal(0, 0,
                                                                              unit.get_carbon_dioxide()))
                unit.epid.set_epid_signal(unit.epid.calculate_epid_signal(0, unit.get_carbon_dioxide()))
                unit.guard.set_guard_water_potential(unit.guard.calculate_guard_water_potential(
                    unit.get_temperature()))
                unit.epid.set_epid_water_potential(unit.epid.calculate_epid_water_potential(
                    unit.get_temperature()))
                unit.set_pore_water_vapor(unit.calculate_pore_water_vapor(ambient_water))
                unit.set_pore_water_potential(unit.calculate_pore_water_potential())

                ws1.cell(i + 1, unit.get_col() + 1).value = unit.get_temperature()
                ws2.cell(i + 1, unit.get_col() + 1).value = unit.get_conductance()

        for i in range(length):
            for unit in array1[i]:
                array2[i][unit.get_col()].set_carbon_dioxide(unit.get_carbon_dioxide)
                array2[i][unit.get_col()].epid.set_epid_water_potential(unit.epid.get_epid_water_potential)

        count += 1
        print(str(count) + "% done.")

        if count == 15:
            for i in range(length):
                for unit in array1[i]:
                    array3[i][unit.get_col()].set_carbon_dioxide(unit.get_carbon_dioxide)
        if count == 16:
            for i in range(length):
                for unit in array1[i]:
                    array4[i][unit.get_col()].set_carbon_dioxide(unit.get_carbon_dioxide)
        if count == 17:
            for i in range(length):
                for unit in array1[i]:
                    array5[i][unit.get_col()].set_carbon_dioxide(unit.get_carbon_dioxide)
        if count == 18:
            for i in range(length):
                for unit in array1[i]:
                    array6[i][unit.get_col()].set_carbon_dioxide(unit.get_carbon_dioxide)
        if count == 19:
            for i in range(length):
                for unit in array1[i]:
                    array7[i][unit.get_col()].set_carbon_dioxide(unit.get_carbon_dioxide)

    for i in range(length):
        for unit in array1[i]:
            unit.set_total_intensity(total_intensity)
            unit.set_blue_intensity(blue_intensity)
            unit.set_absorbed_intensity()

    while count < 25:
        ws1 = temp_file.create_sheet(str(count))
        ws2 = cond_file.create_sheet(str(count))

        for i in range(len(array1)):
            for unit in array1[i]:
                neighbors_carbon.clear()
                neighbors_water.clear()

                if count % 5 == 0:
                    if i != 0 and unit.get_col() != 0 and i != length - 1 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array3[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() != 0 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array3[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[length - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][unit.get_col()].
                                               epid.get_epid_water_potential()())

                    elif unit.get_col() == 0 and i != length - 1 and i != 0:
                        neighbors_carbon.append(array3[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() != 0 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array3[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[0][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[0][unit.get_col()].epid.get_epid_water_potential()())

                    elif unit.get_col() == width - 1 and i != 0 and i != length - 1:
                        neighbors_carbon.append(array3[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() == 0:
                        neighbors_carbon.append(array3[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[length - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][unit.get_col()].
                                               epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() == width - 1:
                        neighbors_carbon.append(array3[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i + 1][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[length - 1][width - 1].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][width - 1].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() == 0:
                        neighbors_carbon.append(array3[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[0][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() == width - 1:
                        neighbors_carbon.append(array3[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[0][width - 1].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[0][width - 1].epid.get_epid_water_potential()())

                elif count % 5 == 1:
                    if i != 0 and unit.get_col() != 0 and i != length - 1 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array4[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() != 0 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array4[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[length - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][unit.get_col()].
                                               epid.get_epid_water_potential()())

                    elif unit.get_col() == 0 and i != length - 1 and i != 0:
                        neighbors_carbon.append(array4[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() != 0 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array4[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[0][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[0][unit.get_col()].epid.get_epid_water_potential()())

                    elif unit.get_col() == width - 1 and i != 0 and i != length - 1:
                        neighbors_carbon.append(array4[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() == 0:
                        neighbors_carbon.append(array4[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[length - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][unit.get_col()].
                                               epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() == width - 1:
                        neighbors_carbon.append(array4[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i + 1][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[length - 1][width - 1].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][width - 1].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() == 0:
                        neighbors_carbon.append(array4[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[0][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() == width - 1:
                        neighbors_carbon.append(array4[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[0][width - 1].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[0][width - 1].epid.get_epid_water_potential()())

                if count % 5 == 2:
                    if i != 0 and unit.get_col() != 0 and i != length - 1 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array5[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() != 0 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array5[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[length - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][unit.get_col()].
                                               epid.get_epid_water_potential()())

                    elif unit.get_col() == 0 and i != length - 1 and i != 0:
                        neighbors_carbon.append(array5[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() != 0 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array5[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[0][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[0][unit.get_col()].epid.get_epid_water_potential()())

                    elif unit.get_col() == width - 1 and i != 0 and i != length - 1:
                        neighbors_carbon.append(array5[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() == 0:
                        neighbors_carbon.append(array5[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[length - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][unit.get_col()].
                                               epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() == width - 1:
                        neighbors_carbon.append(array5[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i + 1][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[length - 1][width - 1].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][width - 1].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() == 0:
                        neighbors_carbon.append(array5[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[0][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() == width - 1:
                        neighbors_carbon.append(array5[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[0][width - 1].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[0][width - 1].epid.get_epid_water_potential()())

                if count % 5 == 3:
                    if i != 0 and unit.get_col() != 0 and i != length - 1 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array6[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() != 0 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array6[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[length - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][unit.get_col()].
                                               epid.get_epid_water_potential()())

                    elif unit.get_col() == 0 and i != length - 1 and i != 0:
                        neighbors_carbon.append(array6[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() != 0 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array6[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[0][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[0][unit.get_col()].epid.get_epid_water_potential()())

                    elif unit.get_col() == width - 1 and i != 0 and i != length - 1:
                        neighbors_carbon.append(array6[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() == 0:
                        neighbors_carbon.append(array6[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[length - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][unit.get_col()].
                                               epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() == width - 1:
                        neighbors_carbon.append(array6[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i + 1][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[length - 1][width - 1].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][width - 1].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() == 0:
                        neighbors_carbon.append(array6[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[0][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() == width - 1:
                        neighbors_carbon.append(array6[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[0][width - 1].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[0][width - 1].epid.get_epid_water_potential()())

                if count % 5 == 4:
                    if i != 0 and unit.get_col() != 0 and i != length - 1 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array7[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() != 0 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array7[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[length - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][unit.get_col()].
                                               epid.get_epid_water_potential()())

                    elif unit.get_col() == 0 and i != length - 1 and i != 0:
                        neighbors_carbon.append(array7[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() != 0 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array7[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[0][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[0][unit.get_col()].epid.get_epid_water_potential()())

                    elif unit.get_col() == width - 1 and i != 0 and i != length - 1:
                        neighbors_carbon.append(array7[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() == 0:
                        neighbors_carbon.append(array7[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[length - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][unit.get_col()].
                                               epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() == width - 1:
                        neighbors_carbon.append(array7[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i + 1][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[length - 1][width - 1].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][width - 1].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() == 0:
                        neighbors_carbon.append(array7[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[0][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() == width - 1:
                        neighbors_carbon.append(array7[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[0][width - 1].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[0][width - 1].epid.get_epid_water_potential()())

                average_carbon = sum(neighbors_carbon) / len(neighbors_carbon)
                average_water = sum(neighbors_water) / len(neighbors_water)
                unit.calculate_next(average_water)
                x, y = fsolve(unit.solve_for_temperature, (296, 30))
                unit.set_temperature(x)
                unit.set_es_water_vapor(y)
                unit.set_es_water_potential(unit.calculate_es_water_potential(ambient_water))
                unit.set_carbon_dioxide(unit.calculate_carbon_dioxide(0, ambient_carbon, average_carbon))
                unit.guard.set_guard_signal(unit.guard.calculate_guard_signal(0, 0,
                                                                              unit.get_carbon_dioxide()))
                unit.epid.set_epid_signal(unit.epid.calculate_epid_signal(0, unit.get_carbon_dioxide()))
                unit.guard.set_guard_water_potential(unit.guard.calculate_guard_water_potential(
                    unit.get_temperature()))
                unit.epid.set_epid_water_potential(unit.epid.calculate_epid_water_potential(
                    unit.get_temperature()))
                unit.set_pore_water_vapor(unit.calculate_pore_water_vapor(ambient_water))
                unit.set_pore_water_potential(unit.calculate_pore_water_potential())

                ws1.cell(i + 1, unit.get_col() + 1).value = unit.get_temperature()
                ws2.cell(i + 1, unit.get_col() + 1).value = unit.get_conductance()

        for i in range(length):
            for unit in array1[i]:
                array2[i][unit.get_col()].epid.set_epid_water_potential(unit.epid.get_epid_water_potential)
                if count % 5 == 0:
                    array3[i][unit.get_col()].set_carbon_dioxide(unit.get_carbon_dioxide)
                elif count % 5 == 1:
                    array4[i][unit.get_col()].set_carbon_dioxide(unit.get_carbon_dioxide)
                elif count % 5 == 2:
                    array5[i][unit.get_col()].set_carbon_dioxide(unit.get_carbon_dioxide)
                elif count % 5 == 3:
                    array6[i][unit.get_col()].set_carbon_dioxide(unit.get_carbon_dioxide)
                else:
                    array7[i][unit.get_col()].set_carbon_dioxide(unit.get_carbon_dioxide)

        count += 1
        print(str(count) + "% done.")

    while count < 100:
        ws1 = temp_file.create_sheet(str(count))
        ws2 = cond_file.create_sheet(str(count))

        for i in range(len(array1)):
            for unit in array1[i]:
                neighbors_carbon.clear()
                neighbors_water.clear()

                if count % 5 == 0:
                    if i != 0 and unit.get_col() != 0 and i != length - 1 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array3[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() != 0 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array3[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[length - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][unit.get_col()].
                                               epid.get_epid_water_potential()())

                    elif unit.get_col() == 0 and i != length - 1 and i != 0:
                        neighbors_carbon.append(array3[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() != 0 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array3[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[0][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[0][unit.get_col()].epid.get_epid_water_potential()())

                    elif unit.get_col() == width - 1 and i != 0 and i != length - 1:
                        neighbors_carbon.append(array3[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() == 0:
                        neighbors_carbon.append(array3[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[length - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][unit.get_col()].
                                               epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() == width - 1:
                        neighbors_carbon.append(array3[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i + 1][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[length - 1][width - 1].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][width - 1].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() == 0:
                        neighbors_carbon.append(array3[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[0][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() == width - 1:
                        neighbors_carbon.append(array3[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array3[0][width - 1].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[0][width - 1].epid.get_epid_water_potential()())

                elif count % 5 == 1:
                    if i != 0 and unit.get_col() != 0 and i != length - 1 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array4[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() != 0 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array4[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[length - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][unit.get_col()].
                                               epid.get_epid_water_potential()())

                    elif unit.get_col() == 0 and i != length - 1 and i != 0:
                        neighbors_carbon.append(array4[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() != 0 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array4[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[0][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[0][unit.get_col()].epid.get_epid_water_potential()())

                    elif unit.get_col() == width - 1 and i != 0 and i != length - 1:
                        neighbors_carbon.append(array4[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() == 0:
                        neighbors_carbon.append(array4[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[length - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][unit.get_col()].
                                               epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() == width - 1:
                        neighbors_carbon.append(array4[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i + 1][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[length - 1][width - 1].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][width - 1].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() == 0:
                        neighbors_carbon.append(array4[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[0][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() == width - 1:
                        neighbors_carbon.append(array4[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array4[0][width - 1].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[0][width - 1].epid.get_epid_water_potential()())

                if count % 5 == 2:
                    if i != 0 and unit.get_col() != 0 and i != length - 1 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array5[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() != 0 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array5[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[length - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][unit.get_col()].
                                               epid.get_epid_water_potential()())

                    elif unit.get_col() == 0 and i != length - 1 and i != 0:
                        neighbors_carbon.append(array5[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() != 0 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array5[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[0][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[0][unit.get_col()].epid.get_epid_water_potential()())

                    elif unit.get_col() == width - 1 and i != 0 and i != length - 1:
                        neighbors_carbon.append(array5[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() == 0:
                        neighbors_carbon.append(array5[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[length - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][unit.get_col()].
                                               epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() == width - 1:
                        neighbors_carbon.append(array5[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i + 1][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[length - 1][width - 1].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][width - 1].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() == 0:
                        neighbors_carbon.append(array5[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[0][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() == width - 1:
                        neighbors_carbon.append(array5[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array5[0][width - 1].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[0][width - 1].epid.get_epid_water_potential()())

                if count % 5 == 3:
                    if i != 0 and unit.get_col() != 0 and i != length - 1 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array6[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() != 0 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array6[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[length - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][unit.get_col()].
                                               epid.get_epid_water_potential()())

                    elif unit.get_col() == 0 and i != length - 1 and i != 0:
                        neighbors_carbon.append(array6[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() != 0 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array6[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[0][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[0][unit.get_col()].epid.get_epid_water_potential()())

                    elif unit.get_col() == width - 1 and i != 0 and i != length - 1:
                        neighbors_carbon.append(array6[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() == 0:
                        neighbors_carbon.append(array6[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[length - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][unit.get_col()].
                                               epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() == width - 1:
                        neighbors_carbon.append(array6[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i + 1][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[length - 1][width - 1].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][width - 1].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() == 0:
                        neighbors_carbon.append(array6[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[0][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() == width - 1:
                        neighbors_carbon.append(array6[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array6[0][width - 1].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[0][width - 1].epid.get_epid_water_potential()())

                if count % 5 == 4:
                    if i != 0 and unit.get_col() != 0 and i != length - 1 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array7[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() != 0 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array7[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[length - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][unit.get_col()].
                                               epid.get_epid_water_potential()())

                    elif unit.get_col() == 0 and i != length - 1 and i != 0:
                        neighbors_carbon.append(array7[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() != 0 and unit.get_col() != width - 1:
                        neighbors_carbon.append(array7[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[0][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[0][unit.get_col()].epid.get_epid_water_potential()())

                    elif unit.get_col() == width - 1 and i != 0 and i != length - 1:
                        neighbors_carbon.append(array7[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() == 0:
                        neighbors_carbon.append(array7[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i + 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[length - 1][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][unit.get_col()].
                                               epid.get_epid_water_potential()())

                    elif i == 0 and unit.get_col() == width - 1:
                        neighbors_carbon.append(array7[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i + 1][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[length - 1][width - 1].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i + 1][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[length - 1][width - 1].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() == 0:
                        neighbors_carbon.append(array7[i][unit.get_col() + 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i][width - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[0][unit.get_col()].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() + 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][width - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[1][unit.get_col()].epid.get_epid_water_potential()())

                    elif i == length - 1 and unit.get_col() == width - 1:
                        neighbors_carbon.append(array7[i][unit.get_col() - 1].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i][0].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[i - 1][unit.get_col()].get_carbon_dioxide()())
                        neighbors_carbon.append(array7[0][width - 1].get_carbon_dioxide()())

                        neighbors_water.append(array2[i][unit.get_col() - 1].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i][0].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[i - 1][unit.get_col()].epid.get_epid_water_potential()())
                        neighbors_water.append(array2[0][width - 1].epid.get_epid_water_potential()())

                average_carbon = sum(neighbors_carbon) / len(neighbors_carbon)
                average_water = sum(neighbors_water) / len(neighbors_water)
                unit.calculate_next(average_water)
                x, y = fsolve(unit.solve_for_temperature, (296, 30))
                unit.set_temperature(x)
                unit.set_es_water_vapor(y)
                unit.set_es_water_potential(unit.calculate_es_water_potential(ambient_water))
                unit.set_carbon_dioxide(unit.calculate_carbon_dioxide(total_intensity, ambient_carbon,
                                                                      average_carbon))
                unit.guard.set_guard_signal(unit.guard.calculate_guard_signal(blue_intensity, total_intensity,
                                                                              unit.get_carbon_dioxide()))
                unit.epid.set_epid_signal(unit.epid.calculate_epid_signal(total_intensity,
                                                                          unit.get_carbon_dioxide()))
                unit.guard.set_guard_water_potential(
                    unit.guard.calculate_guard_water_potential(unit.get_temperature()))
                unit.epid.set_epid_water_potential(unit.epid.calculate_epid_water_potential(
                    unit.get_temperature()))
                unit.set_pore_water_vapor(unit.calculate_pore_water_vapor(ambient_water))
                unit.set_pore_water_potential(unit.calculate_pore_water_potential())

                ws1.cell(i + 1, unit.get_col() + 1).value = unit.get_temperature()
                ws2.cell(i + 1, unit.get_col() + 1).value = unit.get_conductance()

        for i in range(length):
            for unit in array1[i]:
                array2[i][unit.get_col()].epid.set_epid_water_potential(unit.epid.get_epid_water_potential)
                if count % 5 == 0:
                    array3[i][unit.get_col()].set_carbon_dioxide(unit.get_carbon_dioxide)
                elif count % 5 == 1:
                    array4[i][unit.get_col()].set_carbon_dioxide(unit.get_carbon_dioxide)
                elif count % 5 == 2:
                    array5[i][unit.get_col()].set_carbon_dioxide(unit.get_carbon_dioxide)
                elif count % 5 == 3:
                    array6[i][unit.get_col()].set_carbon_dioxide(unit.get_carbon_dioxide)
                else:
                    array7[i][unit.get_col()].set_carbon_dioxide(unit.get_carbon_dioxide)

        count += 1
        print(str(count) + "% done.")

    temp_file.save('C:\\Users\\Undergrunt\\Box\\Model\\Temperature.xlsx')
    cond_file.save('C:\\Users\\Undergrunt\\Box\\Model\\Conductance.xlsx')
