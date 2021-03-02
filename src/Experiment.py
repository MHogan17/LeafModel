from Array import Array
from Environment import Environment
from ArrayPainter import paint_array
from statistics import stdev
import time
from openpyxl import Workbook


def ExperimentFactory(name):
    if name == 'light':
        print('Valid experiment')
        return Light()
    elif name == 'humidity':
        print('Also a valid experiment')
        return Humidity()
    else:
        raise RuntimeError('Invalid experiment type.')


class Experiment:
    def __init__(self):
        self.environment = Environment()
        self.workbook = Workbook()

    def run(self):
        raise(RuntimeError("'run' is not defined for this experiment type"))

    def record_data(self, t, array):
        ws = self.workbook.active

        ws['A1'] = 'Time'
        ws['B1'] = 'Mean Temp'
        ws['C1'] = 'Std dev'
        ws['D1'] = 'Average WUE'
        ws['E1'] = 'Std dev'

        temp = []
        wue = []
        for i in range(len(array)):
            for unit in array[i]:
                temp.append(unit.get_temperature())
                wue.append((self.environment.get_ambient_carbon() - unit.get_carbon_dioxide()) /
                           (unit.get_es_water_vapor() - self.environment.get_ambient_water()))

        avg_temp = sum(temp) / len(temp)
        avg_wue = sum(wue) / len(wue)

        ws['A' + str(t + 2)] = t
        ws['B' + str(t + 2)] = avg_temp
        ws['C' + str(t + 2)] = stdev(temp)
        ws['D' + str(t + 2)] = avg_wue
        ws['E' + str(t + 2)] = stdev(wue)
        self.workbook.save("Data.xlsx")

    def


class Light(Experiment):
    def run(self):
        start = time.time()
        self.environment.set_total_intensity(0)
        self.environment.set_blue_intensity(0)
        array = Array()
        water = Array()
        wheel = {}
        array.randomize()
        t = 0
        while t < 350:
            if t < 7:
                wheel[t] = Array()
                array.calculate_next(self.environment, water, wheel[t % 7], t)
            else:
                array.calculate_next(self.environment, water, wheel[(t - 6) % 7], t)
            paint_array(array, str(t))
            self.record_data(t, array)

            if t == 20:
                self.environment.set_total_intensity(800)
            #if t == 200:
             #   self.environment.set_blue_intensity(5)

            t += 1
        end = time.time()
        runtime = end - start
        print("This simulation took " + str(round(runtime / 60)) + " minutes.")


class Humidity(Experiment):
    def run(self):
        start = time.time()
        self.environment.set_total_intensity(800)
        self.environment.set_blue_intensity(5)
        self.environment.set_ambient_water(20)
        array = Array()
        water = Array()
        wheel = {}
        array.randomize()
        t = 0
        while t < 350:
            if t < 7:
                wheel[t] = Array()
                array.calculate_next(self.environment, water, wheel[t % 7], t)
            else:
                array.calculate_next(self.environment, water, wheel[(t - 6) % 7], t)
            paint_array(array, str(t))
            self.record_data(t, array)
            t += 1
            if t == 20:
                self.environment.set_ambient_water(10)
        end = time.time()
        runtime = end - start
        print("This simulation took " + str(round(runtime / 60)) + " minutes.")